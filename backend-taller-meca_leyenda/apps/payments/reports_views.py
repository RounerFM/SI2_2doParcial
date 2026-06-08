"""
Reportes generales para administración: agregados, series temporales y exportación Excel.
"""
from __future__ import annotations

from datetime import date, timedelta
from decimal import Decimal
from io import BytesIO

from django.db.models import Avg, Count, Sum
from django.db.models.functions import TruncDate
from django.http import HttpResponse
from django.utils import timezone
from rest_framework.decorators import api_view, permission_classes
from rest_framework.response import Response

from apps.assignments.models import Assignment, AssignmentStatus
from apps.incidents.models import Incident, IncidentCycleMetric, IncidentStatus, IncidentType
from apps.payments.models import Payment, PaymentStatus
from apps.users.models import Role, User
from apps.users.permissions import IsAdmin
from apps.workshops.models import Workshop, WorkshopRating

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font
except ImportError:  # pragma: no cover
    Workbook = None  # type: ignore[misc, assignment]


def _parse_dates(request) -> tuple[date, date]:
    today = timezone.localdate()
    raw_to = request.query_params.get('date_to')
    raw_from = request.query_params.get('date_from')
    try:
        date_to = date.fromisoformat(raw_to) if raw_to else today
    except ValueError:
        date_to = today
    try:
        date_from = date.fromisoformat(raw_from) if raw_from else (date_to - timedelta(days=30))
    except ValueError:
        date_from = date_to - timedelta(days=30)
    if date_from > date_to:
        date_from, date_to = date_to, date_from
    return date_from, date_to


def _incident_base_qs(request, date_from: date, date_to: date):
    qs = Incident.objects.filter(
        created_at__date__gte=date_from,
        created_at__date__lte=date_to,
    )
    wid = request.query_params.get('workshop_id')
    if wid and str(wid).isdigit():
        qs = qs.filter(assignments__workshop_id=int(wid)).distinct()
    st = request.query_params.get('incident_status')
    if st:
        qs = qs.filter(status=st)
    it = request.query_params.get('incident_type')
    if it:
        qs = qs.filter(incident_type=it)
    return qs


def _payments_for_incidents(incident_qs, request):
    pq = Payment.objects.filter(
        assignment__incident__in=incident_qs,
        status__in=[PaymentStatus.CLIENT_PAID, PaymentStatus.COMMISSION_SETTLED],
    )
    ps = request.query_params.get('payment_status')
    if ps:
        pq = pq.filter(status=ps)
    return pq


def _decimal_to_float(d: Decimal | None) -> float | None:
    if d is None:
        return None
    return float(d)


def build_reports_payload(request) -> dict:
    date_from, date_to = _parse_dates(request)
    incident_qs = _incident_base_qs(request, date_from, date_to)
    payments_qs = _payments_for_incidents(incident_qs, request)

    total_incidents = incident_qs.count()
    by_status = list(
        incident_qs.values('status')
        .annotate(count=Count('id'))
        .order_by('-count')
    )
    by_type = list(
        incident_qs.values('incident_type')
        .annotate(count=Count('id'))
        .order_by('-count')
    )
    by_day = list(
        incident_qs.annotate(day=TruncDate('created_at'))
        .values('day')
        .annotate(count=Count('id'))
        .order_by('day')
    )
    for row in by_day:
        d = row.get('day')
        if hasattr(d, 'isoformat'):
            row['day'] = d.isoformat()
        elif d is not None:
            row['day'] = str(d)[:10]

    assign_qs = Assignment.objects.filter(incident__in=incident_qs)
    assignments_by_status = list(
        assign_qs.values('status').annotate(count=Count('id')).order_by('-count')
    )

    revenue_agg = payments_qs.aggregate(
        total=Sum('total_amount'),
        commission=Sum('commission_amount'),
        net_workshop=Sum('workshop_net_amount'),
    )
    payments_count = payments_qs.count()

    completed_incidents = incident_qs.filter(status=IncidentStatus.COMPLETED).count()
    cancelled_incidents = incident_qs.filter(status=IncidentStatus.CANCELLED).count()
    active_incidents = incident_qs.exclude(
        status__in=[IncidentStatus.COMPLETED, IncidentStatus.CANCELLED]
    ).count()

    cycle = IncidentCycleMetric.objects.filter(assignment__incident__in=incident_qs).aggregate(
        avg_assign=Avg('seconds_to_assignment'),
        avg_arrival=Avg('seconds_to_arrival'),
        avg_resolution=Avg('seconds_total_resolution'),
        avg_ai_confidence=Avg('ai_confidence'),
    )

    ratings_in_scope = WorkshopRating.objects.filter(
        assignment__incident__in=incident_qs,
    )
    avg_rating = ratings_in_scope.aggregate(a=Avg('score'))['a']

    new_clients = User.objects.filter(
        role=Role.CLIENT,
        date_joined__date__gte=date_from,
        date_joined__date__lte=date_to,
    ).count()

    new_workshops = Workshop.objects.filter(
        created_at__date__gte=date_from,
        created_at__date__lte=date_to,
    ).count()

    verified_workshops_total = Workshop.objects.filter(is_verified=True, is_active=True).count()

    top_workshops = list(
        payments_qs.values(
            'assignment__workshop_id',
            'assignment__workshop__name',
        )
        .annotate(
            payments_count=Count('id'),
            revenue=Sum('total_amount'),
            commission=Sum('commission_amount'),
        )
        .order_by('-commission')[:15]
    )

    from apps.payments.operational_analytics import (
        _compute_sla,
        _compute_unattended,
        _top_geo_zones,
        _top_workshops_efficiency,
    )

    top_workshops_efficiency = _top_workshops_efficiency(assign_qs)
    top_geo_zones = _top_geo_zones(incident_qs)
    sla_extra = _compute_sla(assign_qs)
    unattended_extra = _compute_unattended(incident_qs)

    incident_rows = list(
        incident_qs.select_related('client__user', 'vehicle')
        .order_by('-created_at')[:200]
        .values(
            'id',
            'status',
            'incident_type',
            'priority',
            'created_at',
            'closed_at',
            'ai_confidence',
            'client__user__first_name',
            'client__user__last_name',
            'vehicle__brand',
            'vehicle__model',
        )
    )
    for row in incident_rows:
        fn = row.pop('client__user__first_name', '') or ''
        ln = row.pop('client__user__last_name', '') or ''
        row['client_name'] = (fn + ' ' + ln).strip() or '—'
        row['vehicle_label'] = (
            f"{row.pop('vehicle__brand', '') or ''} {row.pop('vehicle__model', '') or ''}".strip()
            or '—'
        )
        ca = row.get('created_at')
        row['created_at'] = ca.isoformat() if ca else None
        cl = row.get('closed_at')
        row['closed_at'] = cl.isoformat() if cl else None

    payment_rows = list(
        payments_qs.select_related('assignment__workshop', 'assignment__incident__client__user')
        .order_by('-paid_at', '-created_at')[:200]
        .values(
            'id',
            'total_amount',
            'commission_amount',
            'workshop_net_amount',
            'status',
            'paid_at',
            'assignment__incident_id',
            'assignment__workshop__name',
            'assignment__incident__client__user__first_name',
            'assignment__incident__client__user__last_name',
        )
    )
    for row in payment_rows:
        fn = row.pop('assignment__incident__client__user__first_name', '') or ''
        ln = row.pop('assignment__incident__client__user__last_name', '') or ''
        row['client_name'] = (fn + ' ' + ln).strip() or '—'
        pa = row.get('paid_at')
        row['paid_at'] = pa.isoformat() if pa else None
        for k in ('total_amount', 'commission_amount', 'workshop_net_amount'):
            v = row.get(k)
            if v is not None:
                row[k] = str(v)

    filter_options = {
        'incident_status': [{'value': c[0], 'label': c[1]} for c in IncidentStatus.choices],
        'incident_type': [{'value': c[0], 'label': c[1]} for c in IncidentType.choices],
        'payment_status': [{'value': c[0], 'label': c[1]} for c in PaymentStatus.choices],
        'assignment_status': [{'value': c[0], 'label': c[1]} for c in AssignmentStatus.choices],
    }

    return {
        'meta': {
            'date_from': date_from.isoformat(),
            'date_to': date_to.isoformat(),
            'generated_at': timezone.now().isoformat(),
        },
        'filters_applied': {
            'workshop_id': request.query_params.get('workshop_id') or None,
            'incident_status': request.query_params.get('incident_status') or None,
            'incident_type': request.query_params.get('incident_type') or None,
            'payment_status': request.query_params.get('payment_status') or None,
        },
        'filter_options': filter_options,
        'kpis': {
            'incidents_total': total_incidents,
            'incidents_active': active_incidents,
            'incidents_completed': completed_incidents,
            'incidents_cancelled': cancelled_incidents,
            'payments_settled_count': payments_count,
            'revenue_total': str(revenue_agg['total'] or Decimal('0')),
            'commission_total': str(revenue_agg['commission'] or Decimal('0')),
            'workshop_net_total': str(revenue_agg['net_workshop'] or Decimal('0')),
            'resolution_rate_pct': round(
                (completed_incidents / total_incidents * 100.0) if total_incidents else 0.0,
                1,
            ),
            'avg_assignment_seconds': _decimal_to_float(cycle['avg_assign']),
            'avg_arrival_seconds': _decimal_to_float(cycle['avg_arrival']),
            'avg_resolution_seconds': _decimal_to_float(cycle['avg_resolution']),
            'avg_ai_confidence': _decimal_to_float(cycle['avg_ai_confidence']),
            'avg_rating': round(float(avg_rating), 2) if avg_rating is not None else None,
            'new_clients_in_period': new_clients,
            'new_workshops_in_period': new_workshops,
            'verified_workshops_total': verified_workshops_total,
            'sla_compliance_pct': sla_extra['sla_compliance_pct'],
            'sla_cases_measured': sla_extra['sla_cases_measured'],
            'incidents_unattended': unattended_extra['incidents_unattended'],
            'cancellation_rate_pct': unattended_extra['cancellation_rate_pct'],
        },
        'charts': {
            'incidents_by_status': by_status,
            'incidents_by_type': by_type,
            'incidents_by_day': by_day,
            'assignments_by_status': assignments_by_status,
            'top_workshops_efficiency': top_workshops_efficiency,
            'top_geo_zones': top_geo_zones,
        },
        'top_workshops': top_workshops,
        'tables': {
            'recent_incidents': incident_rows,
            'recent_payments': payment_rows,
        },
    }


@api_view(['GET'])
@permission_classes([IsAdmin])
def admin_reports_summary(request):
    return Response(build_reports_payload(request))


@api_view(['GET'])
@permission_classes([IsAdmin])
def admin_operational_dashboard(request):
    from apps.payments.operational_analytics import build_operational_dashboard

    return Response(build_operational_dashboard(request))


def _sheet_kv(ws, pairs: list[tuple[str, str]], title: str, header_font: Font):
    ws.append([title])
    ws['A1'].font = header_font
    ws.append([])
    for k, v in pairs:
        ws.append([k, v])


@api_view(['GET'])
@permission_classes([IsAdmin])
def admin_reports_export_xlsx(request):
    if Workbook is None:
        return Response(
            {'error': 'openpyxl no está instalado. Ejecuta: pip install openpyxl'},
            status=500,
        )

    data = build_reports_payload(request)
    wb = Workbook()
    bold = Font(bold=True)

    # --- Resumen ---
    ws0 = wb.active
    ws0.title = 'Resumen'
    k = data['kpis']
    meta = data['meta']
    pairs = [
        ('Período desde', meta['date_from']),
        ('Período hasta', meta['date_to']),
        ('Generado', meta['generated_at']),
        ('', ''),
        ('Incidentes (total en filtros)', str(k['incidents_total'])),
        ('Incidentes activos', str(k['incidents_active'])),
        ('Incidentes completados', str(k['incidents_completed'])),
        ('Incidentes cancelados', str(k['incidents_cancelled'])),
        ('Tasa resolución %', str(k['resolution_rate_pct'])),
        ('Pagos liquidados (conteo)', str(k['payments_settled_count'])),
        ('Ingresos brutos (cliente)', k['revenue_total']),
        ('Comisión plataforma', k['commission_total']),
        ('Neto talleres', k['workshop_net_total']),
        ('Seg. media hasta asignación', str(k['avg_assignment_seconds'] or '—')),
        ('Seg. media hasta llegada', str(k['avg_arrival_seconds'] or '—')),
        ('Seg. media resolución total', str(k['avg_resolution_seconds'] or '—')),
        ('Confianza IA media', str(k['avg_ai_confidence'] or '—')),
        ('Calificación media (casos con rating)', str(k['avg_rating'] or '—')),
        ('Clientes nuevos en período', str(k['new_clients_in_period'])),
        ('Talleres nuevos en período', str(k['new_workshops_in_period'])),
        ('Talleres verificados activos (total)', str(k['verified_workshops_total'])),
    ]
    _sheet_kv(ws0, pairs, 'Reporte plataforma — resumen', bold)

    def write_table_sheet(name: str, headers: list[str], rows: list[list]):
        ws = wb.create_sheet(title=name[:31])
        ws.append(headers)
        for c in ws[1]:
            c.font = bold
        for row in rows:
            ws.append(row)

    # --- Por estado / tipo (encabezados en español) ---
    write_table_sheet(
        'Incidentes por estado',
        ['Estado', 'Cantidad'],
        [[r['status'], r['count']] for r in data['charts']['incidents_by_status']],
    )
    write_table_sheet(
        'Incidentes por tipo',
        ['Tipo de incidente', 'Cantidad'],
        [[r['incident_type'], r['count']] for r in data['charts']['incidents_by_type']],
    )
    write_table_sheet(
        'Incidentes por día',
        ['Fecha', 'Cantidad'],
        [[r['day'], r['count']] for r in data['charts']['incidents_by_day']],
    )
    write_table_sheet(
        'Asignaciones por estado',
        ['Estado de asignación', 'Cantidad'],
        [[r['status'], r['count']] for r in data['charts']['assignments_by_status']],
    )
    write_table_sheet(
        'Top talleres',
        ['ID taller', 'Nombre del taller', 'Nº de pagos', 'Ingresos brutos', 'Comisión'],
        [
            [
                r['assignment__workshop_id'],
                r['assignment__workshop__name'],
                r['payments_count'],
                str(r['revenue'] or ''),
                str(r['commission'] or ''),
            ]
            for r in data['top_workshops']
        ],
    )

    inc_field_headers = [
        ('ID', 'id'),
        ('Estado', 'status'),
        ('Tipo', 'incident_type'),
        ('Prioridad', 'priority'),
        ('Cliente', 'client_name'),
        ('Vehículo', 'vehicle_label'),
        ('Fecha de creación', 'created_at'),
        ('Fecha de cierre', 'closed_at'),
        ('Confianza IA', 'ai_confidence'),
    ]
    write_table_sheet(
        'Detalle incidentes',
        [h for h, _ in inc_field_headers],
        [
            [row.get(key) for _, key in inc_field_headers]
            for row in data['tables']['recent_incidents']
        ],
    )

    pay_field_headers = [
        ('ID pago', 'id'),
        ('ID incidente', 'assignment__incident_id'),
        ('Taller', 'assignment__workshop__name'),
        ('Cliente', 'client_name'),
        ('Total cobrado', 'total_amount'),
        ('Monto comisión', 'commission_amount'),
        ('Neto taller', 'workshop_net_amount'),
        ('Estado del pago', 'status'),
        ('Fecha de pago', 'paid_at'),
    ]
    write_table_sheet(
        'Detalle pagos',
        [h for h, _ in pay_field_headers],
        [
            [row.get(key) for _, key in pay_field_headers]
            for row in data['tables']['recent_payments']
        ],
    )

    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)

    fname = f"reporte_admin_{data['meta']['date_from']}_{data['meta']['date_to']}.xlsx"
    resp = HttpResponse(
        bio.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    resp['Content-Disposition'] = f'attachment; filename="{fname}"'
    return resp
